import numpy as np
import openpyxl
import json
import random
import math
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
'''
wb=openpyxl.load_workbook('C:\\Users\\47257\\Desktop\\influence_data.xlsx')
sheet=wb['influence_data']
a0=list(sheet.columns)[0]
a1=list(sheet.columns)[4]
b0=list(sheet.columns)[2]
b1=list(sheet.columns)[6]
K={}
for i in (range(len(a0))):
    K[a0[i].value]=b0[i].value
for i in range(len(a0)):
    K[a1[i].value]=b1[i].value
del K["influencer_id"]
del K["follower_id"]
with open('C:\\Users\\47257\\Desktop\\4.json','w')as f1:
    json.dump(K,f1)
'''

wb=openpyxl.load_workbook("C:\\Users\\47257\\Desktop\\data_by_artist.xlsx")
print(type(wb))
print('begin')
sheet=wb['data_by_artist']
tup=list(sheet.columns)[1]
with open("C:\\Users\\47257\\Desktop\\4.json",'r')as f1:
    a0=json.load(f1)
j=2
for i in tup[1:]:
    if str(i.value) in a0.keys():
        sheet[f'R{j}']=a0[str(i.value)]
    j=j+1
wb.save('C:\\Users\\47257\\Desktop\\12.xlsx')



'''
def solve_0(j,sheet):
    i=1  #i是追随了多少人
    while(True):
        k_0=j+i
        if(sheet[f'E{j}'].value==sheet[f'E{k_0}'].value):
            i=i+1
        else:
            break
    temp_0=0
    for t in range(i):
        k_1=j+t
        temp_0=float(sheet[f'I{k_1}'].value)+temp_0
    for t in range(i):
        k_1=j+t
        sheet[f'J{k_1}']=((sheet[f'I{k_1}'].value)/temp_0)
    j=j+i
    return j




wb=openpyxl.load_workbook('C:\\Users\\47257\\Desktop\\k0.xlsx')
sheet=wb['k0']
j=2
while(j<42771):
    j=solve_0(j,sheet)
wb.save('C:\\Users\\47257\\Desktop\\111.xlsx')
'''


'''
wb=openpyxl.load_workbook('C:\\Users\\47257\\Desktop\\blue.xlsx')
wb1=openpyxl.load_workbook('C:\\Users\\47257\\Desktop\\use\\p2.xlsx')
sheet=wb['blue']
sheet1=wb1['Sheet1']
for i in range(100):
    m=i+1
    u=i+2
    A=random.randint(1,1100)
    tup=list(sheet.rows)[A]
    sheet1[f'A{u}']=tup[3].value
    sheet1[f'B{u}']=tup[4].value
    sheet1[f'C{u}']=tup[9].value
    sheet1[f'D{u}']=tup[10].value
wb1.save('C:\\Users\\47257\\Desktop\\use\\f2.xlsx')
'''
'''
A=list([0.70421,0.470375,0.139954511,0.367813046,0.9844])
B=list([0.667098	,0.440596	,0.214136877	,0.411459825	,0.9998])
C=[0.444411,	0.66057	,0.631894	,0.073792197	,0.4752]
D=[0.42082	,0.646972	,0.5996726	,0.073289219	,0.4772]
def solve_0(A,B):
    Q=0
    W=0
    E=0
    for i in range(len(A)):

        Q=A[i]*B[i]+Q
        W=A[i]*A[i]+W
        E=B[i]*B[i]+E
    f=Q/(math.sqrt(W)*math.sqrt(E))
    return f

f=solve_0(A,D)
print(f)
'''






'''
wb=openpyxl.load_workbook('C:\\Users\\47257\\Desktop\\1.xlsx')
blue=wb['Blues']
elc=wb['Electronic']
pop=wb['Pop Rock']
R=wb['R&B;']

def gensol(gen,numt):#输入流派sheet，和遍历目标长度,1925年到2020年
    temp=[]
    st=1925
    j=2
    io=0
    while(j<numt):
        if gen[f'Q{j}'].value==st:
            j=checkdevice(gen,j,temp)
            st=st+1
            io=0
        elif gen[f'Q{j}'].value!=st and io ==0 :
            try:
                yt=temp[-1]
            except IndexError:
                yt=0
            temp.append(yt)
            st=st+1
            io=1
        else:
            temp.append(0)
            st=st+1
    if st!=2021:
        while(len(temp)!=96):
            temp.append(0)
    return temp
def checkdevice(gen,j,temp):
    tp=j
    num=1
    temp_0=0
    while(gen[f'Q{j}'].value==gen[f'Q{j+1}'].value):
        num=num+1
        j=j+1
    j=j+1
    for i in range(num):
        temp_0=temp_0+gen[f'p{tp}'].value
        tp=tp+1
    r=temp_0/num
    temp.append(r)
    return j
def plotgraph_0(x,temp,ax):
    ax.plot(x, temp, linewidth=0.5)
x = np.linspace(1925, 2020, 96)
temp1=np.array(gensol(blue,1180))
temp2=np.array(gensol(elc,1197))
temp3=np.array(gensol(pop,46270))
temp4=np.array(gensol(R,9746))
fig, ax = plt.subplots()
plotgraph_0(x,temp1,ax)
plotgraph_0(x,temp2,ax)
plotgraph_0(x,temp3,ax)
plotgraph_0(x,temp4,ax)
plt.legend(['blue','elc','pop','R'])
ax.xaxis.set_major_locator(ticker.MaxNLocator(30))
plt.show()
'''